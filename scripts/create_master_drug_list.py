__author__ = 'mattdyer'

from optparse import OptionParser
from openpyxl import load_workbook
from openpyxl import Workbook

#store drug information
class Drug:
    ## The constructor
    # @param self The object pointer
    # @param name The drug name
    # @param status The regulatory status
    # @param area The therapeutic area
    # @param target The therapeutic targets
    # @param pathway The pathway target
    # @param patentExp The expiration date of the patent
    # @param flag The no patent found flag
    def __init__(self, name, status, area, target, pathway, patentExp, flag, source, catalog):
        self.__name = name
        self.__cas = ''
        self.__status = status
        self.__area = area
        self.__target = target
        self.__pathway = pathway
        self.__patentExp = patentExp
        self.__flag = flag
        self.__source = source
        self.__catalog = catalog
        self.__interactions = {}

    ## Add an interactor to the drug
    # @param self The object pointer
    # @param protein The protein interactor
    def addInteractor(self, protein):
        self.__interactions[protein] = 1

    ## Add an cas to the drug
    # @param self The object pointer
    # @param cas The cas number
    def addCAS(self, cas):
        self.__cas = cas

    ## Return the na,e
    # @param self The object pointer
    # @returns The drug name
    def getName(self):
        return(self.__name)

    ## Return the CAS
    # @param self The object pointer
    # @returns The drug CAS
    def getCAS(self):
        return(self.__cas)

    ## Return the status
    # @param self The object pointer
    # @returns The drug status
    def getStatus(self):
        return(self.__status)

    ## Return the area
    # @param self The object pointer
    # @returns The drug area
    def getArea(self):
        return(self.__area)

    ## Return the target
    # @param self The object pointer
    # @returns The drug target
    def getTarget(self):
        return(self.__target)

    ## Return the pathway
    # @param self The object pointer
    # @returns The drug pathway
    def getPathway(self):
        return(self.__pathway)

    ## Return the patent expiration date
    # @param self The object pointer
    # @returns The drug patent expiration date
    def getPatentExp(self):
        return(self.__patentExp)

    ## Return the patent flag
    # @param self The object pointer
    # @returns The drug patent flag
    def getFlag(self):
        return(self.__flag)

    ## Return the patent source
    # @param self The object pointer
    # @returns The drug patent source
    def getSource(self):
        return(self.__source)

    ## Return the catalog
    # @param self The object pointer
    # @returns The drug catalog number
    def getCatalog(self):
        return(self.__catalog)

    ## Return the
    # @param self The object pointer
    # @returns The drug
    def getInteractors(self):
        interactors = '; '.join(list(self.__interactions.keys()))
        return(interactors)

#start here when the script is launched
if (__name__ == '__main__'):
    #set up the option parser
    parser = OptionParser()

    #add the options to parse
    parser.add_option('-a', '--unapproved_drugs', dest='unapproved', help='Unapproved Drug spreadsheet')
    parser.add_option('-b', '--unapproved_interactions', dest='unapprovedInteractions', help='Unapproved Interaction spreadsheet')
    parser.add_option('-c', '--unapproved_cas', dest='unapprovedCAS', help='Unapproved CAS spreadsheet')
    parser.add_option('-d', '--approved_drugs', dest='approved', help='Approved drug spreadsheet')
    parser.add_option('-e', '--approved_interactions', dest='approvedInteractions', help='Approved interactions file')
    parser.add_option('-o', '--output', dest='output', help='The output directory')
    (options, args) = parser.parse_args()

    #load the unapproved drug list
    wb = load_workbook(filename=options.unapproved,use_iterators=True)
    ws = wb.get_sheet_by_name('drugs')
    drugsMap = {}

    #loop over each row
    for row in ws.iter_rows():
        drug = Drug(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, '', '')
        drugsMap[row[0].value] = drug

    #load the unapproved CAS map XLS and parse the data to a map
    wb = load_workbook(filename=options.unapprovedCAS,use_iterators=True)
    ws = wb.get_sheet_by_name('Sheet1')

    #loop over each row
    for row in ws.iter_rows():
        #skip rows with CAS numbers
        if not row[1].value == None:
            drugNames = row[0].value
            tokens = drugNames.split('\n')

            #for each drug add to the map
            for token in tokens:
                #see if drug is important
                if token in drugsMap:
                    drugsMap[token].addCAS(row[1].value)

    #now let's load the unapproved interactions
    wb = load_workbook(filename=options.unapprovedInteractions,use_iterators=True)
    ws = wb.get_sheet_by_name('PPIs')

    #loop over each row
    for row in ws.iter_rows():
        #skip rows with no interactions
        if not row[1].value == None:
            drugNames = row[0].value
            tokens = drugNames.split('\n')

            #for each drug add to the map
            for token in tokens:
                #see if drug is important
                if token in drugsMap:
                    drugsMap[token].addInteractor(row[1].value)


    #load the interactions
    file = open(options.approvedInteractions, 'r')
    interactions = {};

    for line in file:
        line = line.strip()
        tokens = line.split('\t')

        if not tokens[0] in interactions:
            interactions[tokens[0]] = tokens[1]
        else:
            interactions[tokens[0]] = '%s; %s' % (interactions[tokens[0]], tokens[1])

    #close the file
    file.close()

    #load approved drugs
    wb = load_workbook(filename=options.approved,use_iterators=True)
    ws = wb.get_sheet_by_name('Sheet1')

    #loop over each row
    for row in ws.iter_rows():
        drug = Drug(row[0].value, 'Approved', row[2].value, '', '', '', '', row[9].value, row[8].value)
        drug.addCAS(row[3].value)

        if row[3].value in interactions:
            drug.addInteractor(interactions[row[3].value])

        drugsMap[row[0].value] = drug

    #now let's print everything out
    wb = Workbook(optimized_write = True)
    ws = wb.create_sheet()
    for drugLabel in drugsMap.keys():
        drug = drugsMap[drugLabel]

        print '%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s' % (drug.getName(), drug.getCAS(), drug.getStatus(), drug.getPatentExp(), drug.getArea(), drug.getTarget(), drug.getInteractors(), '', drug.getSource(), drug.getCatalog(), drug.getFlag())




